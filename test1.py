from openpyxl import load_workbook
import pandas as pd
import numpy as np
from tkinter import *
import matplotlib.pyplot as pl
from tkinter import messagebox

root=Tk()


def onclick(value):
    if value==1:
        def click():#Student report card
            a=enter.get()

            excel_file='abhinav.xlsx'
            df=pd.read_excel(excel_file)
            print(df)
            l=df['Name'].tolist()
            if a not  in l:  
                    x=messagebox.showinfo('Message!' , 'No Student Found')
                    
                                
            else:     
            
                df.set_index("Name",inplace=True)
                b=df['Roll No'][a]
                c=df['Math'][a]
                d=df['Physics'][a]
                e=df['Chem'][a]
                f=df['computer'][a]
                g=df['English'][a]
                print(b,c,d,e,f,g)

                x=df.loc[:,"Math":"English"].sum(axis=1)
                df['Total']=x
                sum1=df['Total'][a]
                print(sum1)
                

                top=Toplevel()
                
                frame=LabelFrame(top , text='STUDENT REPORT', padx=50 , pady=50 , bg='white' )
                frame.pack()
                name1=Label(frame, text='Name:', padx=10 , pady=10 , width=12,
                        height=1, borderwidth=3, relief="groove" , font=("Calibri", 15), anchor="w" ).grid(row=0 , column=0)
                a1=Label(frame, text=a , padx=10 , pady=10 , width=11,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=0 , column=1)

                rollno=Label(frame,text='Roll No:' , padx=10 , pady=10, width=12,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15), anchor="w").grid(row=0 , column=2)
                b1=Label(frame, text=b , padx=10 , pady=10, width=11,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=0 , column=3)

                math_mark=Label(frame,text='Math:' , padx=10 , pady=10 , width=26,
                        height=1, borderwidth=3, relief="groove" , font=("Calibri", 15) , anchor="w" ).grid(row=2 , column=0 , columnspan=2 )
                c1=Label(frame, text=c , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=2 , column=2, columnspan=2)

                phy_mark=Label(frame, text='Physics:' , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15), anchor="w").grid(row=3 , column=0, columnspan=2)
                d1=Label(frame, text=d , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=3 , column=2, columnspan=2)

                chem_mark=Label(frame,text='Chemistry:' , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15), anchor="w").grid(row=4 , column=0, columnspan=2)
                e1=Label(frame, text=e , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)) .grid(row=4 , column=2, columnspan=2)

                comp_mark=Label(frame,text='Computer:' , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15), anchor="w").grid(row=5 , column=0, columnspan=2)
                f1=Label(frame, text=f , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=5 , column=2, columnspan=2)

                eng_mark=Label(frame,text='English:' , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15), anchor="w") .grid(row=6, column=0, columnspan=2)
                g1=Label(frame, text=g , padx=10 , pady=10, width=26,
                        height=1, borderwidth=3, relief="groove", font=("Calibri", 15)).grid(row=6 , column=2, columnspan=2)

                total=Label(frame,text='Total:' + str(sum1) , padx=10 , pady=10, width=55,
                        height=1, borderwidth=3, relief="flat", font=("Calibri", 15) , bg="green" , fg='white') .grid(row=7, column=0, columnspan=4)

        ent_name=Toplevel()
        frame2=LabelFrame(ent_name , text='STUDENT REPORT', padx=25 , pady=25 )
        frame2.pack(expand=True , fill=BOTH)
        enter_name=Label(frame2, text='Enter Name:', font=("Calibri", 15)).pack(expand=True, fill=BOTH)
        enter=Entry(frame2)
        enter.pack(expand=True, fill=BOTH)

        button1=Button(frame2, text='Submit' , command=click , width=10 ,height=1, font=("Calibri", 15)).pack(expand=True, fill=BOTH)
    elif value==2:#input student
        input_name=Toplevel()
        input_name.rowconfigure(0 , weight=1)
        input_name.columnconfigure(0 , weight=1)
        frame1=LabelFrame(input_name , padx=50 , pady=25)
        frame1.pack()
        lb1=Label(frame1  , text='Name:' ,  font=("Calibri", 15), padx=5  , width='10' , anchor='w' , height=2).grid(row=0 , column=0 )
        lb2=Label(frame1  , text='Roll No:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=1 , column=0 )
        lb3=Label(frame1  , text='Math:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=2 , column=0 )
        lb4=Label(frame1  , text='Physics:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=3 , column=0 )
        lb5=Label(frame1  , text='Chemistry:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=4 , column=0 )
        lb6=Label(frame1  , text='computer:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=5 , column=0 )
        lb7=Label(frame1  , text='English:', font=("Calibri", 15), padx=5 , width='10' , anchor='w' , height=2).grid(row=6 , column=0 )

        en1=Entry(frame1)
        en1.grid(row=0 , column=1 )
        en2=Entry(frame1)
        en2.grid(row=1 , column=1 )
        en3=Entry(frame1)
        en3.grid(row=2 , column=1 )
        en4=Entry(frame1)
        en4.grid(row=3 , column=1 )
        en5=Entry(frame1)
        en5.grid(row=4 , column=1 )
        en6=Entry(frame1)
        en6.grid(row=5 , column=1 )
        en7=Entry(frame1)
        en7.grid(row=6 , column=1 )

        def submit():
            re=pd.read_excel('abhinav.xlsx')
            env1=(en1.get())
            env2=int(en2.get())
            env3=int(en3.get())
            env4=int(en4.get())
            env5=int(en5.get())
            env6=int(en6.get())    
            env7=int(en7.get())
            
            
            wrkbk = load_workbook('abhinav.xlsx')
            # to get the active work sheet
            sh = wrkbk.active
            a=sh.max_row
            
            sh.cell(row=a+1 , column=1 , value=env1)
            sh.cell(row=a+1 , column=2 , value=env2)
            sh.cell(row=a+1 , column=3 , value=env3)
            sh.cell(row=a+1 , column=4 , value=env4)
            sh.cell(row=a+1 , column=5 , value=env5)
            sh.cell(row=a+1 , column=6 , value=env6)
            sh.cell(row=a+1 , column=7 , value=env7)
            wrkbk.save('abhinav.xlsx')
            messagebox.showinfo('Message!' , 'Submitted Succesfully')  
            df=pd.read_excel('abhinav.xlsx')
            df=df.sort_values(['Roll No'])
            print(df)
           
        
            
            
        def reset():
            en1.delete(0 , END)
            en2.delete(0 , END)
            en3.delete(0 , END)
            en4.delete(0 , END)
            en5.delete(0 , END)
            en6.delete(0 , END)
            en7.delete(0 , END)

                 
        button1=Button(frame1 , text="Submit",command=submit  , width=10 ,  font=("Calibri", 15) ).grid(row=7 , column=0  )
        button2=Button(frame1 , text="Reset",command=reset , width=10 ,  font=("Calibri", 15)).grid(row=7 , column=1  )
    elif value==3:#students above average
        df = pd.read_excel("abhinav.xlsx")
        print(df)
        wrkbk = load_workbook('abhinav.xlsx')
        sh = wrkbk.active
        maxrow=sh.max_row

        print("Student List\n")
        print(df.loc[:,:"Roll No"],'\n')
        sub=['Math','Physics','Chem','computer','English']
        color=['r','gold','g','b','c']

        a=df.loc[:,"Math":"English"].sum(axis=1)
        df['Total']=a
        df['Percentage']=(a/500)*100

        v=df.sort_values(by='Total',ascending=False)
        v['Rank']=np.arange(1,maxrow)
        y=v['Rank']
        df['Rank']=y 
        print(a)

        total_Mean=df["Total"].mean()
        print(total_Mean)
        mean1=df['Name'].where(df['Total']>total_Mean).dropna()
        df1=list(mean1)
        print(*df1 , sep='\n')
        total_mean1=float(total_Mean)

        top1=Toplevel()
        frame3=LabelFrame(top1 , padx=25 , pady=25, bg='pale green' )
        frame3.pack()
        lbl1=Label(frame3 , text='Students who scored above average:' , font=('Times New Roman' , 15) , bg='green' , padx=10 , pady=10 , width=25).grid(row=0 , column=0)
        avg=Listbox(frame3 , height=15 , width=25  ,borderwidth=3, relief="flat" , bg='pale green' , font=('Calibri' ,15)    )
        avg.grid(row=1 , column=0)
        label1=Label(frame3 , bg='dodger blue' , fg='white', text='Average='+str(total_mean1) , padx=10 , pady=10 , font=('Times New Roman' , 15), width=25).grid(row=2 , column=0)
        for item in df1:
                avg.insert(END ,  item )
    elif value==4:#delete student
        def enter():
                 excel_file='abhinav.xlsx'
                 df=pd.read_excel(excel_file)
                 print(df)
                 a=entry1.get()
                 l=df['Name'].tolist()
                 if a not  in l:  
                    x=messagebox.showinfo('Message!' , 'No Student Found')
                 else:
                        wrkbk = load_workbook("abhinav.xlsx")
                        sheet=wrkbk["Sheet1"]
                        df=pd.read_excel("abhinav.xlsx")
                        print(sheet)
                        
                        b=df.index[df['Name']==a].tolist()
                        length=len(b)
                        i=0
                        while i <length :
                                c=b[i]
                                print(c)
                                i+=1
                        
                        sheet.delete_rows(c+2 ,1)
                        wrkbk.save('abhinav.xlsx')
                        df=pd.read_excel('abhinav.xlsx')
                        df=df.sort_values(['Roll No'])
                        print(df)
                        y=messagebox.showinfo('Message!' , 'Deleted Successfully!!!')
                
        
        top2=Toplevel()
        frame4=LabelFrame(top2 , padx=25 , pady=25)
        frame4.pack()
        lbl2=Label(frame4 , text='ENTER STUDENT NAME' , fg='red' ,font=("Calibri", 15)).grid(row=0 , column=0 )
        entry1=Entry(frame4 , width=50)
        entry1.grid(row=1 , column=0)
        but1=Button(frame4 , text='Enter' ,command=enter   , font=("Calibri", 15)).grid(row=2 , column=0)
    elif value==5:#statistics
        
        df=pd.read_excel('abhinav.xlsx')
        a=df.loc[:,"Math":"English"].sum(axis=1)
        df['Total']=a
        df['Percentage']=(a/500)*100
        b=df['Percentage']

        pl.figure(figsize=(13,10))
        plot1 = pl.subplot2grid((3, 3), (0, 0), colspan=3 )
        plot2 = pl.subplot2grid((3, 3), (1, 0),  colspan=3 , rowspan=2)

        range=(0, 100)
        bins=10
        plot1.hist(b, bins, range, color = 'green',
                histtype = 'bar', rwidth = 0.8)
        pl.xlabel('Percentage')
        pl.ylabel('No. of Students')
        pl.title('Marks vs. No. of Students')

        print(df.loc[:,:"English"])
        plot2.bar(df.Name,df.Percentage,color='orange'  )
        pl.ylim(df.Percentage.min()-3,df.Percentage.max()+1)
        pl.xlabel('Student name')
        pl.ylabel('Percentage')
        pl.xticks(rotation=45)
        pl.show()

    elif value==6:#failed students
            
        top6=Toplevel()

        df=pd.read_excel('abhinav.xlsx')
        print(df)
        b=df['Name'].where(df['Math']<35).dropna()
        b1=b.tolist()
        print(b1)
        c=df['Name'].where(df['Physics']<35).dropna()
        c1=c.tolist()
        print(c1)
        d=df['Name'].where(df['Chem']<35).dropna()
        d1=d.tolist()
        e=df['Name'].where(df['computer']<35).dropna()
        e1=e.tolist()
        f=df['Name'].where(df['English']<35).dropna()
        f1=f.tolist()
        print(d1,e1,f1)
        frame=LabelFrame(top6 , padx=35 , pady=35 )
        frame.pack()
        lbl11=Label(frame , padx=10 , pady=10 ,font=('Arial Bold',15) , anchor='w', text='Students who failed in:' , width=42,
                                height=1, borderwidth=3, relief="groove").grid(row=0 , column=0, columnspan=2)
        lbl2=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , anchor='w', text='Math:' , width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=1 , column=0)
        lbl3=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text='Physics:', anchor='w', width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=2 , column=0)
        lbl4=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text='Chemistry:', anchor='w', width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=3 , column=0)
        lbl5=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text='Computer:', anchor='w', width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=4 , column=0)
        lbl6=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text='English:', anchor="w", width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=5 , column=0)
        lbl1=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text=c1, width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=2 , column=1)
        lbl7=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text=b1, width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=1 , column=1)
        lbl8=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text=d1, width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=3 , column=1)
        lbl9=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text=e1, width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=4 , column=1)
        lbl10=Label(frame , padx=10 , pady=10 , font=('Arial Bold',15) , text=f1, width=20,
                                height=1, borderwidth=3, relief="groove").grid(row=5 , column=1)
        lbl11=Label(frame , padx=10 , pady=10 ,font=('Arial Bold',15) , text='Pass mark= 35' , width=42,
                                height=1, borderwidth=3, relief="groove", bg='green' , fg='white').grid(row=6 , column=0, columnspan=2)


        

r=IntVar()
frame=LabelFrame(root , text='STUDENT REPORT' , padx=10 , pady=10 , bg='white') 
frame.pack(expand = True, fill = BOTH)

Radiobutton(frame, variable=r , value=1 ,text='Student Report Card', anchor='w' , font=("Arial Bold", 20),bg='dodger blue' , width=30).pack(anchor='w',expand = True, fill = BOTH)
Radiobutton(frame, variable=r , value=2 , text='Input new student' , anchor='w' , font=("Arial Bold", 20),bg='red', width=30).pack(anchor='w',expand = True, fill = BOTH)
Radiobutton(frame, variable=r , value=3 , text='View list of students above average' , anchor='w' , font=("Arial Bold", 20),bg='yellow', width=30).pack(anchor='w',expand = True, fill = BOTH)
Radiobutton(frame, variable=r , value=4 , text='Delete Student' , anchor='w' , font=("Arial Bold", 20),bg='green', width=30).pack(anchor='w',expand = True, fill = BOTH)
Radiobutton(frame, variable=r , value=5 , text='Statistics' , anchor='w' , font=("Arial Bold", 20),bg='purple', width=30).pack(anchor='w',expand = True, fill = BOTH)
Radiobutton(frame, variable=r , value=6 , text='Failed Students' , anchor='w' , font=("Arial Bold", 20),bg='aqua', width=30).pack(anchor='w',expand = True, fill = BOTH)

bt1=Button(frame , text='Submit' , font=("Calibri", 15), command=lambda: onclick(r.get())).pack(expand = True, fill = BOTH)

root.mainloop()

